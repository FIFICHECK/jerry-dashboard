import openpyxl, json, os, glob
from datetime import datetime, timedelta

# --- Config ---
# Auto-detect latest SKU List XLSX (no need to update path manually each run)
downloads = '/home/snkwok/Downloads'
xlsx_files = sorted(glob.glob(os.path.join(downloads, 'SKU List*.xlsx')), key=os.path.getmtime)
xlsx_path = xlsx_files[-1] if xlsx_files else '/home/snkwok/Downloads/SKU List (1).xlsx'

repo_dir = os.environ.get('JERRY_REPO_DIR', '/home/snkwok/jerry-dashboard')

# === Jerry store whitelist (security: only Jerry's stores' data goes public) ===
# Source of truth: contactData in index.html (Jerry's 119 stores).
# SKU data from the XLSX covers the WHOLE platform — filter to Jerry's stores
# so non-Jerry merchants' GMV/qty never leaks to the public repo.
def load_jerry_whitelist():
    import re as _re
    html_path = os.path.join(repo_dir, 'index.html')
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            content = f.read()
        m = _re.search(r'const contactData = (\{.*?\});', content, _re.DOTALL)
        if not m:
            print('WARN: contactData not found in index.html — whitelist empty', flush=True)
            return set()
        import json as _json
        return set(_json.loads(m.group(1)).keys())
    except Exception as e:
        print(f'WARN: could not load whitelist: {e}', flush=True)
        return set()

JERRY_STORES = load_jerry_whitelist()
print(f'Jerry store whitelist: {len(JERRY_STORES)} stores', flush=True)

# Compute month from yesterday (MTD)
yesterday = datetime.now() - timedelta(days=1)
month = yesterday.strftime('%Y-%m')
today_str = datetime.now().strftime('%Y-%m-%d')
print(f"XLSX: {xlsx_path}")
print(f"Month: {month}, Date: {today_str}")

# === Load historical name lookup ===
full_json_path = os.path.join(repo_dir, 'sku_data_full.json')
with open(full_json_path, 'r') as f:
    full_data = json.load(f)
name_lookup = {}
for entry in full_data:
    sc = entry['sc']
    sn = entry.get('sn', '')
    if sn and sn.strip():
        name_lookup[sc] = sn
print(f'Historical SKUs with names: {len(name_lookup)}')

# === Parse XLSX ===
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active
headers = [cell.value for cell in ws[1]]
print(f"Headers: {headers}")

col_map = {}
for i, h in enumerate(headers):
    col_map[h] = i

sku_idx = col_map['primary_sku_code']
name_idx = col_map['primary_sku_name_chi']
gmv_idx = col_map['GMV']
qty_idx = col_map['Qty']

print(f"Column indices: sku={sku_idx}, name={name_idx}, gmv={gmv_idx}, qty={qty_idx}")

rows = []
empty_names = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    sc = row[sku_idx]
    if not sc:
        continue
    # SECURITY: skip non-Jerry stores (only Jerry's store data goes to the public repo)
    if JERRY_STORES and str(sc).split('_')[0] not in JERRY_STORES:
        continue
    gmv = float(row[gmv_idx] or 0)
    qty = int(row[qty_idx] or 0)
    sn = (row[name_idx] or '').strip()
    if not sn:
        empty_names += 1
    rows.append({'sc': sc, 'sn': sn, 'gmv': gmv, 'qty': qty, 'm': month})

rows.sort(key=lambda x: -x['gmv'])

filled_from_history = 0
still_empty = 0
for r in rows:
    if not r['sn']:
        if r['sc'] in name_lookup:
            r['sn'] = name_lookup[r['sc']]
            filled_from_history += 1
        else:
            still_empty += 1

print(f'Total: {len(rows)} SKUs, Empty in XLSX: {empty_names}')
print(f'Filled from history: {filled_from_history}, Still empty: {still_empty}')
print(f'Total GMV: {sum(r["gmv"] for r in rows):.2f}')

sku_data_json = json.dumps(rows, ensure_ascii=False, separators=(',', ':'))

# === Update sku_data_full.json ===
old_count = len(full_data)
# SECURITY: also drop any non-Jerry store entries lingering in history
if JERRY_STORES:
    full_data = [e for e in full_data if e['sc'].split('_')[0] in JERRY_STORES]
full_data = [e for e in full_data if e['m'] != month]
removed = old_count - len(full_data)
print(f'Removed {removed} old entries for {month}')

full_data.extend(rows)
full_data.sort(key=lambda x: (-int(x['m'].replace('-', '')), -x['gmv']))

with open(full_json_path, 'w') as f:
    json.dump(full_data, f, ensure_ascii=False)
print(f'sku_data_full.json: {len(full_data)} entries')

# === Update index.html ===
html_path = os.path.join(repo_dir, 'index.html')

with open(html_path, 'r') as f:
    content = f.read()

lines = content.split('\n')
sku_updated = False
for i, line in enumerate(lines):
    stripped = line.strip()
    if stripped.startswith('const skuData = [') and stripped.endswith('];'):
        indent = line[:len(line) - len(line.lstrip())]
        lines[i] = indent + 'const skuData = ' + sku_data_json + ';'
        print(f'skuData updated at line {i+1} ({len(rows)} entries)')
        sku_updated = True
        break

if not sku_updated:
    print('ERROR: skuData line not found!')
    exit(1)

content = '\n'.join(lines)

with open(html_path, 'w') as f:
    f.write(content)

# Date badge update (binary mode)
with open(html_path, 'rb') as f:
    raw = f.read()

date_prefix = b'\xf0\x9f\x93\x85 \xe6\x95\xb8\xe6\x93\x9a\xe4\xb8\x8b\xe8\xbc\x89\xe6\x97\xa5\xe6\x9c\x9f: '
if date_prefix in raw:
    start = raw.find(date_prefix) + len(date_prefix)
    old_date = raw[start:start+10].decode('utf-8')
    raw = raw.replace(date_prefix + old_date.encode(), date_prefix + today_str.encode())
    print(f'Date badge: {old_date} → {today_str}')
else:
    print('WARNING: Date badge prefix not found')

with open(html_path, 'wb') as f:
    f.write(raw)

# === Verify Sessions 1-6 structural integrity ===
with open(html_path, 'rb') as f:
    final_raw = f.read()
checks = {
    'targetStoreMonthData': b'const targetStoreMonthData' in final_raw,
    'months array': b'const months' in final_raw,
    'storeMonthly': b'const storeMonthly' in final_raw,
    'contactData': b'const contactData' in final_raw,
    'Target_GMV': b'Target_GMV' in final_raw,
}
all_ok = all(checks.values())
sku_count = final_raw.count(b'"sc":')
for name, ok in checks.items():
    print(f'  {name}: {"✅" if ok else "❌"}')
print(f'SKU count in file: {sku_count}')
print(f'Session integrity: {"PASS" if all_ok else "FAIL"}')

print(f'\n=== SUMMARY ===')
print(f'XLSX: {len(rows)} SKUs parsed')
print(f'sku_data_full.json: {len(full_data)} entries total')
print(f'Date badge updated to {today_str}')
print(f'SKU count in file: {sku_count}')
print(f'Session integrity: {"PASS" if all_ok else "FAIL"}')
